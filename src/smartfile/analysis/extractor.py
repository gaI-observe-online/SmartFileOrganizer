"""Content extraction from various file types."""

import logging
from pathlib import Path
from typing import Dict, Optional

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    from docx import Document
except ImportError:
    Document = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    from PIL import Image
    import pytesseract
except ImportError:
    Image = None
    pytesseract = None

try:
    import eml_parser
except ImportError:
    eml_parser = None


logger = logging.getLogger(__name__)


class ContentExtractor:
    """Extract text content from various file types."""
    
    MAX_CONTENT_LENGTH = 1000  # Extract first 1000 chars
    
    def extract(self, file_path: Path) -> Dict[str, any]:
        """Extract content from file.
        
        Args:
            file_path: Path to file
            
        Returns:
            Dictionary with extracted data:
            - content: Text content (first 1000 chars)
            - metadata: File metadata
            - entities: Extracted entities (names, dates, etc.)
            - doc_type: Detected document type
        """
        result = {
            "content": "",
            "metadata": {},
            "entities": [],
            "doc_type": "unknown"
        }
        
        if not file_path.exists():
            return result
        
        suffix = file_path.suffix.lower()
        
        try:
            if suffix == '.pdf':
                result.update(self._extract_pdf(file_path))
            elif suffix in ['.doc', '.docx']:
                result.update(self._extract_word(file_path))
            elif suffix in ['.xlsx', '.xls']:
                result.update(self._extract_excel(file_path))
            elif suffix == '.csv':
                result.update(self._extract_csv(file_path))
            elif suffix in ['.txt', '.md']:
                result.update(self._extract_text(file_path))
            elif suffix in ['.jpg', '.jpeg', '.png', '.gif', '.bmp']:
                result.update(self._extract_image(file_path))
            elif suffix == '.eml':
                result.update(self._extract_email(file_path))
            else:
                # Try as text file
                result.update(self._extract_text(file_path))
        except Exception as e:
            logger.warning(f"Failed to extract content from {file_path}: {e}")
        
        return result
    
    def _extract_pdf(self, file_path: Path) -> Dict[str, any]:
        """Extract content from PDF file."""
        result = {"content": "", "metadata": {}, "doc_type": "pdf"}
        
        if not pdfplumber:
            logger.warning("pdfplumber not available, skipping PDF extraction")
            return result
        
        try:
            with pdfplumber.open(file_path) as pdf:
                # Extract metadata
                result["metadata"] = {
                    "pages": len(pdf.pages),
                    "info": pdf.metadata or {}
                }
                
                # Extract text from first few pages
                text_parts = []
                for page in pdf.pages[:3]:  # First 3 pages
                    text = page.extract_text()
                    if text:
                        text_parts.append(text)
                    if len(''.join(text_parts)) >= self.MAX_CONTENT_LENGTH:
                        break
                
                result["content"] = ''.join(text_parts)[:self.MAX_CONTENT_LENGTH]
                
                # Detect document type based on content
                content_lower = result["content"].lower()
                if any(word in content_lower for word in ['invoice', 'bill', 'payment']):
                    result["doc_type"] = "invoice"
                elif any(word in content_lower for word in ['contract', 'agreement', 'terms']):
                    result["doc_type"] = "contract"
                elif any(word in content_lower for word in ['resume', 'curriculum vitae', 'cv']):
                    result["doc_type"] = "resume"
                
        except Exception as e:
            logger.warning(f"Error extracting PDF {file_path}: {e}")
        
        return result
    
    def _extract_word(self, file_path: Path) -> Dict[str, any]:
        """Extract content from Word document."""
        result = {"content": "", "metadata": {}, "doc_type": "document"}
        
        if not Document:
            logger.warning("python-docx not available, skipping Word extraction")
            return result
        
        try:
            doc = Document(file_path)
            
            # Extract metadata
            core_props = doc.core_properties
            result["metadata"] = {
                "author": core_props.author or "",
                "title": core_props.title or "",
                "created": str(core_props.created) if core_props.created else ""
            }
            
            # Extract text from paragraphs
            text_parts = []
            for para in doc.paragraphs:
                text_parts.append(para.text)
                if len(''.join(text_parts)) >= self.MAX_CONTENT_LENGTH:
                    break
            
            result["content"] = '\n'.join(text_parts)[:self.MAX_CONTENT_LENGTH]
            
        except Exception as e:
            logger.warning(f"Error extracting Word document {file_path}: {e}")
        
        return result
    
    def _extract_excel(self, file_path: Path) -> Dict[str, any]:
        """Extract content from Excel file."""
        result = {"content": "", "metadata": {}, "doc_type": "spreadsheet"}
        
        if not load_workbook:
            logger.warning("openpyxl not available, skipping Excel extraction")
            return result
        
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            
            result["metadata"] = {
                "sheets": wb.sheetnames,
                "sheet_count": len(wb.sheetnames)
            }
            
            # Extract data from first sheet
            sheet = wb.active
            text_parts = []
            
            for row in sheet.iter_rows(max_row=20, values_only=True):
                row_text = ' | '.join(str(cell) if cell is not None else '' for cell in row)
                text_parts.append(row_text)
                if len('\n'.join(text_parts)) >= self.MAX_CONTENT_LENGTH:
                    break
            
            result["content"] = '\n'.join(text_parts)[:self.MAX_CONTENT_LENGTH]
            
            wb.close()
            
        except Exception as e:
            logger.warning(f"Error extracting Excel file {file_path}: {e}")
        
        return result
    
    def _extract_csv(self, file_path: Path) -> Dict[str, any]:
        """Extract content from CSV file."""
        result = {"content": "", "metadata": {}, "doc_type": "spreadsheet"}
        
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                lines = []
                for i, line in enumerate(f):
                    if i >= 20:  # First 20 lines
                        break
                    lines.append(line.strip())
                
                result["content"] = '\n'.join(lines)[:self.MAX_CONTENT_LENGTH]
                
        except Exception as e:
            logger.warning(f"Error extracting CSV file {file_path}: {e}")
        
        return result
    
    def _extract_text(self, file_path: Path) -> Dict[str, any]:
        """Extract content from text file."""
        result = {"content": "", "metadata": {}, "doc_type": "text"}
        
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                result["content"] = f.read(self.MAX_CONTENT_LENGTH)
                
        except Exception as e:
            logger.warning(f"Error extracting text file {file_path}: {e}")
        
        return result
    
    def _extract_image(self, file_path: Path) -> Dict[str, any]:
        """Extract text from image using OCR."""
        result = {"content": "", "metadata": {}, "doc_type": "image"}
        
        if not Image or not pytesseract:
            logger.warning("PIL or pytesseract not available, skipping OCR")
            return result
        
        try:
            img = Image.open(file_path)
            
            result["metadata"] = {
                "size": img.size,
                "format": img.format,
                "mode": img.mode
            }
            
            # Perform OCR
            try:
                text = pytesseract.image_to_string(img)
                result["content"] = text[:self.MAX_CONTENT_LENGTH]
            except Exception as e:
                logger.debug(f"OCR failed for {file_path}: {e}")
            
            img.close()
            
        except Exception as e:
            logger.warning(f"Error extracting image {file_path}: {e}")
        
        return result
    
    def _extract_email(self, file_path: Path) -> Dict[str, any]:
        """Extract content from email (.eml) file."""
        result = {"content": "", "metadata": {}, "doc_type": "email"}
        
        if not eml_parser:
            logger.warning("eml-parser not available, skipping email extraction")
            return result
        
        try:
            with open(file_path, 'rb') as f:
                raw_email = f.read()
            
            ep = eml_parser.EmlParser()
            parsed = ep.decode_email_bytes(raw_email)
            
            result["metadata"] = {
                "from": parsed.get('header', {}).get('from', ''),
                "to": parsed.get('header', {}).get('to', ''),
                "subject": parsed.get('header', {}).get('subject', ''),
                "date": parsed.get('header', {}).get('date', '')
            }
            
            # Extract body
            body = parsed.get('body', [])
            if body:
                result["content"] = body[0].get('content', '')[:self.MAX_CONTENT_LENGTH]
            
        except Exception as e:
            logger.warning(f"Error extracting email {file_path}: {e}")
        
        return result
